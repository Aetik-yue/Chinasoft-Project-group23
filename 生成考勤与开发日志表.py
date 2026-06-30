# -*- coding: utf-8 -*-
"""
生成项目组成员考勤记录和开发日志的 Excel 表格
- 考勤日期：6.29 ~ 7.13（仅工作日）
- 上午 / 下午分开签到签退
- 成员：严浩睿（组长）、马沛霖、郑良锋、陈心澄、王田煜
- 负责模块 / 任务内容 留白
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta


# ============= 样式定义 =============
HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
SUBHEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="333333")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="1F4E78")
NORMAL_FONT = Font(name="微软雅黑", size=10, color="333333")
HINT_FONT = Font(name="微软雅黑", size=10, italic=True, color="808080")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_header_style(ws, row, col_start, col_end, height=28):
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = height


def apply_body_style(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        fill = ALT_FILL if (r - row_start) % 2 == 1 else None
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if fill:
                cell.fill = fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============= 成员信息（负责模块留空） =============
members = [
    # (序号, 姓名, 学号, 角色, 负责模块, 联系方式)
    (1, "严浩睿", "", "组长", "", ""),
    (2, "马沛霖", "", "组员", "", ""),
    (3, "郑良锋", "", "组员", "", ""),
    (4, "陈心澄", "", "组员", "", ""),
    (5, "王田煜", "", "组员", "", ""),
]
member_names = [m[1] for m in members]


# ============= 考勤日期：6.29 ~ 7.13（含周末） =============
def build_all_days(start: date, end: date):
    """返回 [start, end] 区间内所有日期（不再排除周末）"""
    days = []
    cur = start
    while cur <= end:
        days.append(cur)
        cur += timedelta(days=1)
    return days


START_DATE = date(2026, 6, 29)
END_DATE = date(2026, 7, 13)
all_days = build_all_days(START_DATE, END_DATE)


# ============= 工作簿初始化 =============
wb = Workbook()

# ---------------- 使用说明（放在最前） ----------------
ws_help = wb.active
ws_help.title = "使用说明"

ws_help.merge_cells("A1:B1")
ws_help["A1"] = "智慧烟感项目 · 考勤与开发日志表 使用说明"
ws_help["A1"].font = TITLE_FONT
ws_help["A1"].alignment = CENTER
ws_help.row_dimensions[1].height = 32

help_rows = [
    ("工作簿构成", "本文件包含 5 个工作表：使用说明、项目成员信息、考勤记录、开发日志、任务清单。"),
    ("考勤日期范围",
     f"{START_DATE.strftime('%Y-%m-%d')} ~ {END_DATE.strftime('%Y-%m-%d')}"
     f"，共 {len(all_days)} 天（含周末）。"),
    ("推荐填写顺序",
     "1) 先在「项目成员信息」表填写学号、模块、联系方式；\n"
     "2) 在「考勤记录」表按日逐人填入上午/下午的签到签退时间，工作时长会自动计算；\n"
     "3) 在「开发日志」表按日逐人记录今日任务与次日计划。\n"
     "4) 周末/休息日可不填时间，只在备注里写「休息」即可。"),
    ("考勤记录表",
     "• 上午签到 / 上午签退 / 下午签到 / 下午签退 都使用 HH:MM 格式\n"
     "• 出勤状态：出勤 / 迟到 / 早退 / 请假 / 旷工 / 补卡\n"
     "• 工作时长（小时）由公式 =IFERROR(((上午签退-上午签到)+(下午签退-下午签到))*24,2) 自动计算\n"
     "• 如果只上半天，请把不参与的时间段留空，公式会忽略空值"),
    ("开发日志表",
     "• 负责模块、今日任务、遇到的问题、次日计划 等内容目前已留白，请按需填写\n"
     "• 完成情况：已完成 / 进行中 / 阻塞 / 未开始\n"
     "• 进度按 0~100 整数填写\n"
     "• 遇到的问题与次日计划要具体，避免空泛"),
    ("任务清单表",
     "• 用于登记项目整体任务，与每日开发日志互补\n"
     "• 优先级下拉：高 / 中 / 低\n"
     "• 状态下拉：未开始 / 进行中 / 已完成 / 阻塞\n"
     "• 负责人请填写「项目成员信息」表中的成员姓名\n"
     "• 完成度按 0~100 整数填写；开始/截止日期使用 yyyy-mm-dd"),
    ("筛选与排序", "三张数据表（考勤记录、开发日志、任务清单）都已开启筛选器，可按姓名、日期、状态等筛选统计。"),
    ("冻结窗口", "考勤记录、开发日志、任务清单表均冻结到第 5 行，滚动时表头保持可见。"),
    ("提交方式", "每周将本 Excel 文件提交到 Git 仓库 文档/ 目录下，命名建议：考勤与开发日志_第N周.xlsx"),
]

for i, (k, v) in enumerate(help_rows, start=3):
    ws_help.cell(row=i, column=1, value=k).font = SUBHEADER_FONT
    ws_help.cell(row=i, column=1).alignment = CENTER
    ws_help.cell(row=i, column=1).fill = SUBHEADER_FILL
    ws_help.cell(row=i, column=1).border = BORDER
    ws_help.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws_help.cell(row=i, column=2).alignment = LEFT
    ws_help.cell(row=i, column=2).border = BORDER
    ws_help.row_dimensions[i].height = max(30, 18 * (v.count("\n") + 1))

set_col_widths(ws_help, [20, 95])


# ---------------- 项目成员信息 ----------------
ws_members = wb.create_sheet("项目成员信息")

ws_members.merge_cells("A1:F1")
ws_members["A1"] = "智慧烟感项目 · 项目组成员信息表"
ws_members["A1"].font = TITLE_FONT
ws_members["A1"].alignment = CENTER
ws_members.row_dimensions[1].height = 32

ws_members["A2"] = "项目名称：智慧烟感系统"
ws_members["B2"] = "项目组：第23组"
ws_members["D2"] = "填表日期："
ws_members["E2"] = date.today().strftime("%Y-%m-%d")
for c in ("A2", "B2", "D2", "E2"):
    ws_members[c].font = SUBHEADER_FONT if c in ("A2", "D2") else NORMAL_FONT
ws_members.row_dimensions[2].height = 22

member_headers = ["序号", "姓名", "学号/工号", "角色", "负责模块", "联系方式"]
for col, h in enumerate(member_headers, start=1):
    ws_members.cell(row=4, column=col, value=h)
apply_header_style(ws_members, 4, 1, len(member_headers))

for i, row in enumerate(members, start=5):
    for j, v in enumerate(row, start=1):
        ws_members.cell(row=i, column=j, value=v)
apply_body_style(ws_members, 5, 5 + len(members) - 1, 1, len(member_headers))
set_col_widths(ws_members, [6, 14, 14, 10, 22, 22])
ws_members.freeze_panes = "A5"


# ---------------- 考勤记录（上午 / 下午） ----------------
ws_att = wb.create_sheet("考勤记录")

ws_att.merge_cells("A1:J1")
ws_att["A1"] = "智慧烟感项目 · 项目组成员考勤记录表"
ws_att["A1"].font = TITLE_FONT
ws_att["A1"].alignment = CENTER
ws_att.row_dimensions[1].height = 32

ws_att.merge_cells("A2:J2")
ws_att["A2"] = ("考勤日期范围："
                f"{START_DATE.strftime('%Y-%m-%d')} ~ {END_DATE.strftime('%Y-%m-%d')}"
                f"（共 {len(all_days)} 天，含周末）。"
                "上午/下午分别签到签退，工作时长由公式自动计算；"
                "每天的 5 行结束之后有一行空白分隔行，方便区分不同日期。")
ws_att["A2"].font = HINT_FONT
ws_att["A2"].alignment = LEFT
ws_att.row_dimensions[2].height = 30

att_headers = [
    "序号", "日期", "星期", "姓名",
    "上午签到", "上午签退", "下午签到", "下午签退",
    "工作时长(h)", "出勤状态/备注"
]
for col, h in enumerate(att_headers, start=1):
    ws_att.cell(row=4, column=col, value=h)
apply_header_style(ws_att, 4, 1, len(att_headers), height=30)

WEEKDAY_CN = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

# 生成数据：每天 5 名成员一行，每日之间加一行空白分隔
start_row = 5
idx = 1
r = start_row
last_data_row = r - 1
for dt in all_days:
    for name in member_names:
        ws_att.cell(row=r, column=1, value=idx)
        c2 = ws_att.cell(row=r, column=2, value=dt)
        c2.number_format = "yyyy-mm-dd"
        ws_att.cell(row=r, column=3, value=WEEKDAY_CN[dt.weekday()])
        ws_att.cell(row=r, column=4, value=name)
        for col in (5, 6, 7, 8):
            ws_att.cell(row=r, column=col, value="")
        ws_att.cell(
            row=r, column=9,
            value=(
                f'=IFERROR(ROUND('
                f'((IF(E{r}="",0,TIMEVALUE(F{r})-TIMEVALUE(E{r})))+'
                f'(IF(G{r}="",0,TIMEVALUE(H{r})-TIMEVALUE(G{r}))))*24,2),"")'
            )
        )
        ws_att.cell(row=r, column=10, value="")
        last_data_row = r
        idx += 1
        r += 1
    # 每天 5 行之后插入一个空白分隔行
    for col in range(1, len(att_headers) + 1):
        sep = ws_att.cell(row=r, column=col, value=None)
        sep.fill = PatternFill("solid", fgColor="FFFFFF")
        sep.border = Border()
    ws_att.row_dimensions[r].height = 8
    r += 1

end_row = r - 1
apply_body_style(ws_att, start_row, last_data_row, 1, len(att_headers))
set_col_widths(ws_att, [6, 12, 10, 12, 11, 11, 11, 11, 12, 22])

ws_att.auto_filter.ref = f"A4:J{last_data_row}"  # 筛选只覆盖数据行
ws_att.freeze_panes = "A5"


# ---------------- 开发日志（留白） ----------------
ws_log = wb.create_sheet("开发日志")

ws_log.merge_cells("A1:K1")
ws_log["A1"] = "智慧烟感项目 · 开发日志记录表"
ws_log["A1"].font = TITLE_FONT
ws_log["A1"].alignment = CENTER
ws_log.row_dimensions[1].height = 32

ws_log.merge_cells("A2:K2")
ws_log["A2"] = ("填写说明：完成情况填写：已完成、进行中、阻塞、未开始；遇到的问题写具体原因与解决方案；"
                "次日计划要明确可交付物。负责模块与具体任务内容由各成员按实际情况填写；"
                "每天的 5 行结束之后有一行空白分隔行，方便区分不同日期。")
ws_log["A2"].font = HINT_FONT
ws_log["A2"].alignment = LEFT
ws_log.row_dimensions[2].height = 30

log_headers = [
    "序号", "日期", "星期", "姓名", "负责模块", "今日任务", "完成情况",
    "进度(%)", "工时(h)", "遇到的问题", "次日计划"
]
for col, h in enumerate(log_headers, start=1):
    ws_log.cell(row=4, column=col, value=h)
apply_header_style(ws_log, 4, 1, len(log_headers), height=30)

# 生成数据：每天 5 名成员一行，每日之间加一行空白分隔
start_row = 5
idx = 1
r = start_row
last_data_row = r - 1
for dt in all_days:
    for name in member_names:
        ws_log.cell(row=r, column=1, value=idx)
        c2 = ws_log.cell(row=r, column=2, value=dt)
        c2.number_format = "yyyy-mm-dd"
        ws_log.cell(row=r, column=3, value=WEEKDAY_CN[dt.weekday()])
        ws_log.cell(row=r, column=4, value=name)
        for col in range(5, len(log_headers) + 1):
            ws_log.cell(row=r, column=col, value="")
        ws_log.cell(row=r, column=8).number_format = "0"  # 进度整数
        last_data_row = r
        idx += 1
        r += 1
    # 每天 5 行之后插入一个空白分隔行
    for col in range(1, len(log_headers) + 1):
        sep = ws_log.cell(row=r, column=col, value=None)
        sep.fill = PatternFill("solid", fgColor="FFFFFF")
        sep.border = Border()
    ws_log.row_dimensions[r].height = 8
    r += 1

end_row = r - 1
apply_body_style(ws_log, start_row, last_data_row, 1, len(log_headers))
set_col_widths(ws_log, [6, 12, 10, 12, 14, 28, 10, 9, 9, 28, 28])

ws_log.auto_filter.ref = f"A4:K{last_data_row}"  # 筛选只覆盖数据行
ws_log.freeze_panes = "A5"


# ---------------- 任务清单 ----------------
ws_task = wb.create_sheet("任务清单")

ws_task.merge_cells("A1:K1")
ws_task["A1"] = "智慧烟感项目 · 任务清单"
ws_task["A1"].font = TITLE_FONT
ws_task["A1"].alignment = CENTER
ws_task.row_dimensions[1].height = 32

ws_task.merge_cells("A2:K2")
ws_task["A2"] = ("填写说明：用于登记项目整体任务，与每日开发日志互补。"
                "优先级下拉 高/中/低；状态下拉 未开始/进行中/已完成/阻塞；"
                "负责人填写「项目成员信息」表中的成员姓名；完成度 0~100 整数。")
ws_task["A2"].font = HINT_FONT
ws_task["A2"].alignment = LEFT
ws_task.row_dimensions[2].height = 30

task_headers = [
    "序号", "任务名称", "任务描述", "负责人", "协作人",
    "优先级", "状态", "开始日期", "截止日期", "完成度(%)", "备注"
]
for col, h in enumerate(task_headers, start=1):
    ws_task.cell(row=4, column=col, value=h)
apply_header_style(ws_task, 4, 1, len(task_headers), height=30)

# 预置 30 个空白行（仅序号，其余留白）
TASK_ROWS = 30
start_row = 5
for i in range(TASK_ROWS):
    r = start_row + i
    ws_task.cell(row=r, column=1, value=i + 1)
    for col in range(2, len(task_headers) + 1):
        ws_task.cell(row=r, column=col, value="")
    # 日期列格式
    ws_task.cell(row=r, column=8).number_format = "yyyy-mm-dd"
    ws_task.cell(row=r, column=9).number_format = "yyyy-mm-dd"
    ws_task.cell(row=r, column=10).number_format = "0"  # 完成度整数

end_row = start_row + TASK_ROWS - 1
apply_body_style(ws_task, start_row, end_row, 1, len(task_headers))
set_col_widths(ws_task, [6, 22, 32, 12, 12, 10, 12, 12, 12, 11, 22])

# 数据验证：优先级（F 列）
dv_pri = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
dv_pri.add(f"F{start_row}:F{end_row}")
ws_task.add_data_validation(dv_pri)

# 数据验证：状态（G 列）
dv_stat = DataValidation(type="list", formula1='"未开始,进行中,已完成,阻塞"', allow_blank=True)
dv_stat.add(f"G{start_row}:G{end_row}")
ws_task.add_data_validation(dv_stat)

ws_task.auto_filter.ref = f"A4:K{end_row}"
ws_task.freeze_panes = "A5"


# ============= 保存 =============
output_path = r"E:\Chinasoft-Project-group23\文档\项目组考勤与开发日志.xlsx"
wb.save(output_path)
print(f"Excel 文件已生成：{output_path}")
print(f"包含工作表：{wb.sheetnames}")
print(f"考勤日期范围：{START_DATE} ~ {END_DATE}，共 {len(all_days)} 天（含周末）")
print(f"成员：{member_names}")
total_data = len(all_days) * len(members)
print(f"考勤记录数据行：{total_data}（{len(all_days)} 天 × {len(members)} 人），含 {len(all_days)} 个空行分隔")
print(f"开发日志数据行：{total_data}（{len(all_days)} 天 × {len(members)} 人），含 {len(all_days)} 个空行分隔")
